import discord
from discord.ext import commands
from dotenv import load_dotenv
import os
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
import requests

scope = ['https://www.googleapis.com/auth/spreadsheets', 
         'https://www.googleapis.com/auth/drive', 
         'https://www.googleapis.com/auth/drive.file', 
         'https://spreadsheets.google.com/feeds']

credentials = ServiceAccountCredentials.from_json_keyfile_name("google_keys.json", scope)
gspread_client = gspread.authorize(credentials)

Database = gspread_client.open("EC Database").worksheet("Database")

list_usn = Database.col_values(Database.find("USN").col)
list_backlog = Database.col_values(Database.find("Backlog").col)
list_placed = Database.col_values(Database.find("Placement Status").col)
list_usn.pop(0)
list_placed.pop(0)
list_backlog.pop(0)

list_usn_backlog = []
list_open_dream = []
list_dream = []
list_unplaced = []

for i in range(len(list_usn)):
    if list_backlog[i] == "Yes":
        list_usn_backlog.append(list_usn[i])

for i in range(len(list_usn)):
    if list_placed[i] == "Open Dream":
        list_open_dream.append(list_usn[i])
    if list_placed[i] == "Dream":
        list_dream.append(list_usn[i])
    if list_placed[i] == "Unplaced":
        list_unplaced.append(list_usn[i])

load_dotenv()
token = os.getenv("token")

discord_client = commands.Bot(command_prefix='-', help_command = None, case_insensitive = True)

@discord_client.event
async def on_ready():
    await discord_client.change_presence(activity = discord.Game('-help'))
    print('Logged in as {0.user}'.format(discord_client))

@discord_client.group(invoke_without_command=True)
async def help(ctx):
    embed = discord.Embed(title = 'Help', description = 'List of bot commands', color = discord.Color.green())
    embed.add_field(name = 'About Student Database', value = 'backlog, placed, unplaced, opendream, dream', inline = False)
    embed.add_field(name = 'Verify', value = 'verify, cutoff', inline = False)
    embed.set_footer(text = 'For more information on a command try -help <command>, ex: -help placed')
    await ctx.send(embed = embed)

@help.command()
async def backlog(ctx):
    embed = discord.Embed(title = "Backlog", description = 'Returns a list of USNs of students with backlog', color = discord.Color.green())
    embed.add_field(name = '**Syntax**', value = '-backlog')
    await ctx.send(embed = embed)

@help.command()
async def placed(ctx):
    embed = discord.Embed(title = "Placed", description = 'Returns a list of USNs of placed students', color = discord.Color.blue())
    embed.add_field(name = '**Syntax**', value = '-placed')
    await ctx.send(embed = embed)

@help.command()
async def unplaced(ctx):
    embed = discord.Embed(title = "Unplaced", description = 'Returns a list of USNs of unplaced students', color = discord.Color.red())
    embed.add_field(name = '**Syntax**', value = '-unplaced')
    await ctx.send(embed = embed)

@help.command()
async def opendream(ctx):
    embed = discord.Embed(title = "Open Dream", description = 'Returns a list of USNs of students placed in Open Dream companies', color = discord.Color.blue())
    embed.add_field(name = '**Syntax**', value = '-opendream')
    await ctx.send(embed = embed)

@help.command()
async def dream(ctx):
    embed = discord.Embed(title = "Dream", description = 'Returns a list of USNs of students placed in Dream companies', color = discord.Color.purple())
    embed.add_field(name = '**Syntax**', value = '-dream')
    await ctx.send(embed = embed)

@help.command()
async def verify(ctx):
    embed = discord.Embed(title = "Verify", description = 'Returns a list of USNs of students who are already placed in a open dream or a dream company. Also returns a list of USNs of students who have backlogs and have applied for the given company.', color = discord.Color.green())
    embed.add_field(name = '**Syntax**', value = '-verify <attachment> Ex: -verify (attach a .xlsx file)')
    await ctx.send(embed = embed)

@help.command()
async def cutoff(ctx):
    embed = discord.Embed(title = "Cutoff", description = 'Returns a list of USNs of students who are unplaced and are meeting the companies cutoff requirements, also returns a list of students placed in Dream companies who are eligible.', color = discord.Color.green())
    embed.add_field(name = '**Syntax**', value = '-cutoff <cgpa> Ex: -cutoff 7.5')
    await ctx.send(embed = embed)

@discord_client.command(aliases = ['hi', 'hey'])
async def hello(ctx):
    await ctx.send('Haaai ;)')

@discord_client.command()
async def backlog(ctx):
    await ctx.send(list_usn_backlog)

@discord_client.command()
async def opendream(ctx):
    await ctx.send(list_open_dream)

@discord_client.command()
async def dream(ctx):
    await ctx.send(list_dream)
   
@discord_client.command() 
async def unplaced(ctx):
    await ctx.send(list_unplaced)

@discord_client.command()
async def placed(ctx):
    await ctx.send('Open Dream: {} \n\n Dream: {}'.format(list_open_dream, list_dream))

@discord_client.command()
async def verify(ctx):
    await ctx.send('Attach the file you want to verify.')

    # def check(message):
    #     return message.author == ctx.author and bool(message.attachments)
    # await ctx.send('Attach the file you want to verify.')
    # msg = await discord_client.wait_for('message', check=check)
    # attachment_url = msg.attachments[0].url

    def check(message):
        attachments = message.attachments
        if len(attachments) == 0:
            return False
        attachment = attachments[0]
        return attachment.filename.endswith('.xlsx')

    msg = await discord_client.wait_for('message', check=check)
    attachment_url = msg.attachments[0].url

    r = requests.get(attachment_url, allow_redirects=True)
    open('verify.xlsx', 'wb').write(r.content)

    workbook = openpyxl.load_workbook('verify.xlsx')
    worksheet = workbook.active
    list_verify_usn = []
    max_col = worksheet.max_column
    max_row = worksheet.max_row
    for i in range(1, max_col + 1):
        cell = worksheet.cell(row = 1, column = i)
        if cell.value == "USN":
            col_value = i
            break      
    for i in range(2, max_row + 1):
        list_verify_usn.append(worksheet.cell(row = i, column = col_value).value)
    
    list_remove_backlog = [x for x in list_usn_backlog if x in list_verify_usn]
    list_already_opendream = [x for x in list_open_dream if x in list_verify_usn]
    list_already_dream = [x for x in list_dream if x in list_verify_usn] 
    await ctx.send('Already placed in Open Dream: {}'.format(list_already_opendream))
    await ctx.send('Already placed in Dream: {}'.format(list_already_dream))
    await ctx.send('Students with backlog: {}'.format(list_remove_backlog))

# @discord_client.event
# async def on_command_error(ctx, error):
#     if isinstance(error, commands.MissingRequiredArgument):
#         await ctx.send('Wrong usage of command.')    

@discord_client.command()
async def cutoff(ctx, amount: float):
    list_cutoff = []
    list_temp_dream = []
    list_cgpa = Database.col_values(Database.find("CGPA").col)
    list_cgpa.pop(0)
    for i in range(len(list_usn)):
        if (list_cgpa[i] == ''):
            continue
        if float(list_cgpa[i]) > amount and list_placed[i] == 'Unplaced' and list_backlog[i] != 'Yes':
            list_cutoff.append(list_usn[i])
        if float(list_cgpa[i]) > amount and list_placed[i] == 'Dream' and list_backlog[i] != 'Yes':
            list_temp_dream.append(list_usn[i])
    await ctx.send('List of students that can apply:\n{}\nTotal number of eligible students are: {}\n\nList of students who are in dream company and are eligible: {}\nNumber of eligible students placed in Dream company are: {}'.format(list_cutoff, len(list_cutoff), list_temp_dream, len(list_temp_dream)))

discord_client.run(token)